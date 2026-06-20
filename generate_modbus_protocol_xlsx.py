#!/usr/bin/env python3
"""
生成 Modbus TCP 服务器通信协议 xlsx 文件
直接从协议 XML 文件和 EMS tcp_cmd 配置读取，不依赖运行时的 dev_list
每个设备独占一个工作表，包含 FC03(HR) 数据点表
"""

import sys, os, json, xml.etree.ElementTree as ET
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

import config

# ==================== 协议文件 → 设备名映射 ====================
# 格式: (设备名, 中文名, 协议文件路径)
PROTOCOL_DEVICES = [
    ('ems',        'EMS',        None),  # EMS 使用 tcp_cmd.json
    ('pcs1',       'PCS1',       config.EPCS15_AM_COMMUNICATION_FILEPATH),
    ('bms_uhome',   'BMS',        config.BMS_UHOME_COMMUNICATION_FILEPATH),
    ('dtsd3366',   '储能电表',    config.DTSD_3366D_COMMUNICATION_FILEPATH),
    ('ac_wea1610', '空调',       config.WEA1610_FILEPATH),
    ('dehumidifier','除湿机',     None),  # 除湿机协议文件未在config中找到独立XML
]

# 允许用户扩展
_extra = getattr(config, 'EXTRA_PROTOCOL_DEVICES', None)
if _extra:
    PROTOCOL_DEVICES.extend(_extra)

# ==================== 样式 ====================
HDR_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HDR_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
C_ALIGN = Alignment(horizontal='center', vertical='center')
L_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ALT_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')


def parse_xml_registers(xml_path):
    """解析 XML 协议文件，返回有序的寄存器列表 [{name, address, datatype, mag, offset, unit}]"""
    if not xml_path or not os.path.exists(xml_path):
        return []
    tree = ET.parse(xml_path)
    root = tree.getroot()
    regs = []
    fc03 = root.find('function_code03')
    fc04 = root.find('function_code04')
    if fc03 is not None:
        for hr in fc03.findall('hRegister'):
            regs.append({
                'name': hr.get('name', ''),
                'address': int(hr.get('address', 0)),
                'datatype': hr.get('datatype', 'UINT16'),
                'mag': float(hr.get('mag', 1)),
                'offset': float(hr.get('offset', 0)),
                'unit': hr.get('unit', ''),
            })
    if fc04 is not None:
        for hr in fc04.findall('iRegister'):
            regs.append({
                'name': hr.get('name', ''),
                'address': int(hr.get('address', 0)),
                'datatype': hr.get('datatype', 'UINT16'),
                'mag': float(hr.get('mag', 1)),
                'offset': float(hr.get('offset', 0)),
                'unit': hr.get('unit', ''),
            })
    return regs


def load_ems_tcp_cmd():
    """从 ems_tcp_cmd.json 加载 EMS 的 TCP 命令映射"""
    path = getattr(config, 'EMS_TCP_CMD_FILEPATH', './cfg/ems_tcp_cmd.json')
    if not os.path.exists(path):
        return {}
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def generate_xlsx(output_path='modbus_tcp_protocol.xlsx'):
    wb = Workbook()
    wb.remove(wb.active)
    print("[INFO] 正在从协议文件收集寄存器信息...\n")

    # ===== 汇总表 =====
    sw = wb.create_sheet('设备汇总', 0)
    sw.merge_cells('A1:E1')
    sw['A1'].value = 'Modbus TCP 服务器设备汇总'
    sw['A1'].font = Font(name='Arial', size=14, bold=True)
    sw['A1'].alignment = Alignment(horizontal='center')
    for ci, h in enumerate(['序号', '设备名称', '中文名称', '从站ID', '数据点数量'], 1):
        c = sw.cell(row=3, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER
    for ci, w in enumerate([6, 22, 16, 10, 12], 1):
        sw.column_dimensions[get_column_letter(ci)].width = w
    sw.freeze_panes = 'A4'
    sr = 4

    # 加载 EMS tcp_cmd
    ems_tcp_cmd = load_ems_tcp_cmd()

    slave_id = 1
    for dev_name, chinese_name, proto_path in PROTOCOL_DEVICES:
        is_ems = (dev_name == 'ems')
        registers = []

        if is_ems:
            # EMS: 地址 0 = 远程使能，其余从 tcp_cmd.json 读取
            registers.append({
                'name': '远程使能',
                'address': 0, 'tcp_addr': 0,
                'datatype': 'BOOL', 'mag': 1, 'offset': 0, 'unit': '',
            })
            for key, cfg in ems_tcp_cmd.items():
                registers.append({
                    'name': cfg.get('desc', key),
                    'address': cfg.get('tcp_addr', 0),
                    'datatype': cfg.get('datatype', 'UINT16'),
                    'mag': float(cfg.get('mag', 1)),
                    'offset': float(cfg.get('offset', 0)),
                    'unit': cfg.get('unit', ''),
                })
        elif proto_path:
            # 其他设备: 从 XML 读取，地址 0 = 在线状态
            registers.append({
                'name': '在线状态',
                'address': 0, 'tcp_addr': 0,
                'datatype': 'UINT16', 'mag': 1, 'offset': 0, 'unit': '',
            })
            xml_regs = parse_xml_registers(proto_path)
            # 数据点从地址 1 开始
            current_tcp = 1
            for r in xml_regs:
                r['tcp_addr'] = current_tcp
                reg_count = 2 if r['datatype'] in ('INT32', 'UINT32') else 1
                current_tcp += reg_count
                r['address'] = r['tcp_addr']
            registers.extend(xml_regs)
        else:
            print(f"  [SKIP] {dev_name}: 无协议文件")
            continue

        if not registers:
            print(f"  [SKIP] {dev_name}: 无寄存器")
            continue

        # 按地址排序
        registers.sort(key=lambda r: r.get('address', 0))

        print(f"  [{dev_name}] {chinese_name} (ID={slave_id}): {len(registers)} 个数据点")

        # ===== 设备工作表 =====
        ws = wb.create_sheet(title=dev_name[:31])
        ws.merge_cells('A1:J1')
        t = ws['A1']
        t.value = f"设备: {chinese_name} ({dev_name})  |  TCP 从站 ID: {slave_id}"
        t.font = Font(name='Arial', size=14, bold=True, color='1F3864')
        t.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 28

        headers = ['序号', 'TCP地址', 'RTU地址', '功能码', '名称', '数据类型', '倍率', '偏移', '单位', '说明']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN; c.border = BORDER
        ws.row_dimensions[3].height = 22

        for ri, reg in enumerate(registers, start=1):
            row = ri + 3
            addr = reg.get('address', 0)
            ws.cell(row=row, column=1, value=ri).alignment = C_ALIGN
            ws.cell(row=row, column=2, value=addr).alignment = C_ALIGN
            ws.cell(row=row, column=3, value=f"0x{addr:04X}").alignment = C_ALIGN
            ws.cell(row=row, column=4, value='03/04').alignment = C_ALIGN
            ws.cell(row=row, column=5, value=reg['name']).alignment = L_ALIGN
            ws.cell(row=row, column=6, value=reg['datatype']).alignment = C_ALIGN
            ws.cell(row=row, column=7, value=reg['mag']).alignment = C_ALIGN
            ws.cell(row=row, column=8, value=reg['offset']).alignment = C_ALIGN
            ws.cell(row=row, column=9, value=reg.get('unit', '')).alignment = C_ALIGN
            ws.cell(row=row, column=10, value='').alignment = L_ALIGN
            for c in range(1, 11):
                ws.cell(row=row, column=c).border = BORDER
            if ri % 2 == 0:
                for c in range(1, 11):
                    ws.cell(row=row, column=c).fill = ALT_FILL

        for ci, w in enumerate([5, 9, 9, 7, 32, 9, 7, 7, 8, 20], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A4'

        # 汇总行
        for ci, v in enumerate([slave_id, dev_name, chinese_name, slave_id, len(registers)], 1):
            c = sw.cell(row=sr, column=ci, value=v)
            c.alignment = C_ALIGN; c.border = BORDER
        sr += 1
        slave_id += 1

    wb.save(output_path)
    print(f"\n✓ 协议文档: {os.path.abspath(output_path)}")
    print(f"  共 {slave_id - 1} 个设备工作表 + 1 个汇总表")


if __name__ == '__main__':
    output = sys.argv[1] if len(sys.argv) > 1 else 'modbus_tcp_protocol.xlsx'
    generate_xlsx(output)
