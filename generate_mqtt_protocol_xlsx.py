#!/usr/bin/env python3
"""
生成 MQTT 通信协议 xlsx 文件
包含: 每设备发布数据点表 + 控制映射表 + ACK/消息格式说明
"""
import sys, os, json, xml.etree.ElementTree as ET
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

import config

_BASE = os.path.dirname(os.path.abspath(__file__))

# ==================== 样式 ====================
HF = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HFL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
HFP = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
HFC = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
HA = Alignment(horizontal='center', vertical='center', wrap_text=True)
CA = Alignment(horizontal='center', vertical='center')
LA = Alignment(horizontal='left', vertical='center', wrap_text=True)
BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
AF = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')


def load_json(path):
    if not os.path.exists(path): return {}
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


# 加载配置
MQTT_CFG    = load_json(os.path.join(_BASE, 'cfg', 'mqtt_config.json'))
CTRL_MAP    = {k: v for k, v in load_json(os.path.join(_BASE, 'cfg', 'mqtt_cmd_mapping.json')).items() if not k.startswith('_')}
DEV_OFFSETS = MQTT_CFG.get('device_offsets', {})
CTRL_ZH     = load_json(os.path.join(_BASE, 'cfg', 'control_command_zh.json'))
EMS_DATA_DICT     = load_json(os.path.join(_BASE, 'cfg', 'ems_data_dict.json'))

SN = MQTT_CFG.get('sn', 'xxx')
BROKER = MQTT_CFG.get('broker', 'N/A')
PORT = MQTT_CFG.get('port', 1883)

PROTOCOL_DEVICES = [
    ('ems',        'EMS',        None),
    ('pcs1',       'PCS1',       config.EPCS15_AM_COMMUNICATION_FILEPATH),
    ('bms_uhome',  'BMS',        config.BMS_UHOME_COMMUNICATION_FILEPATH),
    ('dtsd3366',   '储能电表',    config.DTSD_3366D_COMMUNICATION_FILEPATH),
    ('ac_wea1610', '空调',       config.WEA1610_FILEPATH),
    ('dehumidifier','除湿机',     None),
]


def parse_xml(xml_path):
    if not xml_path or not os.path.exists(xml_path): return []
    tree = ET.parse(xml_path)
    regs = []
    for tag in ('function_code03', 'function_code04'):
        node = tree.find(tag)
        if node is None: continue
        for hr in (node.findall('hRegister') or node.findall('iRegister') or []):
            regs.append({
                'name': hr.get('name', ''),
                'datatype': hr.get('datatype', 'UINT16'),
                'unit': hr.get('unit', ''),
            })
    return regs


def get_ctrl_zh(dev_name, cmd_key):
    """从 control_command_zh.json 查找命令的中文名"""
    dev_map = CTRL_ZH.get(dev_name, {})
    return dev_map.get(cmd_key, {}).get('zh', cmd_key)


def generate_xlsx(output_path='mqtt_protocol.xlsx'):
    wb = Workbook()
    wb.remove(wb.active)

    # ===== 汇总 =====
    sw = wb.create_sheet('汇总', 0)
    sw.merge_cells('A1:F1')
    sw['A1'].value = 'MQTT 通信协议 —— EMS 能量管理系统'
    sw['A1'].font = Font(name='Arial', size=14, bold=True)
    sw['A1'].alignment = Alignment(horizontal='center')
    sw.merge_cells('A2:F2')
    sw['A2'].value = (f'Broker: {BROKER}:{PORT} | SN: {SN} | '
                      f'发布: cloud/push/{SN}/real | 控制: cloud/action/{SN}/control | ACK: cloud/action/{SN}/control/ack')
    sw['A2'].font = Font(size=9, color='666666')
    for ci, h in enumerate(['工作表', '类型', '设备名', '中文名', '数据点', '说明'], 1):
        c = sw.cell(row=4, column=ci, value=h)
        c.font = HF; c.fill = HFL; c.alignment = HA; c.border = BD
    for ci, w in enumerate([22, 8, 12, 10, 8, 45], 1):
        sw.column_dimensions[get_column_letter(ci)].width = w
    sw.freeze_panes = 'A5'
    sr = 5

    # ===== 消息格式示例 =====
    ws_fmt = wb.create_sheet('消息格式')
    ws_fmt.merge_cells('A1:B1'); ws_fmt['A1'].value = 'MQTT 消息格式说明'; ws_fmt['A1'].font = Font(size=14, bold=True)

    topics = [
        ('发布主题 (push)', f'cloud/push/{SN}/real'),
        ('控制主题 (action)', f'cloud/action/{SN}/control'),
        ('ACK 回复主题', f'cloud/action/{SN}/control/ack'),
    ]
    for ri, (label, topic) in enumerate(topics, start=3):
        ws_fmt.cell(row=ri, column=1, value=label).font = Font(bold=True)
        ws_fmt.cell(row=ri, column=2, value=topic).font = Font(name='Consolas')
    ws_fmt.column_dimensions['A'].width = 20; ws_fmt.column_dimensions['B'].width = 45

    ws_fmt.cell(row=7, column=1, value='发布数据格式').font = Font(bold=True, size=12)
    ws_fmt.merge_cells('A8:B8')
    ws_fmt['A8'].value = '{"type":"changed","online_devices":["pcs1","ems"],"datetime":"...","timestamp":...,"change_count":1,"data":[{"id":5100,"value":296.7}]}'
    ws_fmt['A8'].font = Font(name='Consolas', size=9)
    ws_fmt.cell(row=9, column=1, value='type: "full"=全量(60s) / "changed"=增量  |  alarm_data 字段随告警附带')
    ws_fmt['A9'].font = Font(size=9, color='666666')

    ws_fmt.cell(row=11, column=1, value='定时模式格式 (id=200)').font = Font(bold=True)
    ws_fmt.merge_cells('A12:B16')
    ws_fmt['A12'].value = ('{"id":200,"value":{"chargeTimeList":[{'
                           '"startTime":"00:00","endTime":"16:18","power":14,'
                           '"weekday":["0","0","1","0","1","0","0"]},...],'
                           '"dischargeTimeList":[{'
                           '"startTime":"08:00","endTime":"20:00","power":-5,'
                           '"weekday":["1","1","1","1","1","1","1"]},...]}}')
    ws_fmt['A12'].font = Font(name='Consolas', size=9); ws_fmt['A12'].alignment = Alignment(wrap_text=True)

    ws_fmt.cell(row=18, column=1, value='需求响应格式 (id=300)').font = Font(bold=True)
    ws_fmt.merge_cells('A19:B20')
    ws_fmt['A19'].value = ('{"id":300,"value":[{'
                           '"activePower":-3,"endDatetime":"2026-05-29 00:00",'
                           '"reactivePower":0,"startDatetime":"2026-05-13 00:00"},...]}')
    ws_fmt['A19'].font = Font(name='Consolas', size=9); ws_fmt['A19'].alignment = Alignment(wrap_text=True)

    ws_fmt.cell(row=22, column=1, value='控制消息格式').font = Font(bold=True, size=12)
    ws_fmt.merge_cells('A23:B23')
    ws_fmt['A23'].value = '{"cmd_id":"0","value":1,"sn":"xxx","code":"xxx"}'
    ws_fmt['A23'].font = Font(name='Consolas', size=10)
    ws_fmt.cell(row=24, column=1, value='ACK 回复格式').font = Font(bold=True)
    ws_fmt.merge_cells('A25:B25')
    ws_fmt['A25'].value = ('{"type":"ack","sn":"xxx","code":"xxx","datetime":"...","status":1,"error_msg":"","payload":{...}}  |  '
                           'status=1 成功, status=0 失败(SN/CODE不匹配)')
    ws_fmt['A25'].font = Font(name='Consolas', size=9); ws_fmt['A25'].alignment = Alignment(wrap_text=True)
    for r in range(3, 26):
        ws_fmt.row_dimensions[r].height = max(ws_fmt.row_dimensions[r].height or 15, 18 if r not in (8, 12, 19, 24) else 55)

    sw.cell(sr, 1, '消息格式').alignment = CA; sw.cell(sr, 2, '说明').alignment = CA
    sw.cell(sr, 3, '-').alignment = CA; sw.cell(sr, 4, '-').alignment = CA
    sw.cell(sr, 5, 6).alignment = CA; sw.cell(sr, 6, '主题定义、消息格式、定时/需求响应示例').alignment = LA
    for c in range(1, 7): sw.cell(sr, c).border = BD
    sr += 1

    # ===== 控制映射表 =====
    ws = wb.create_sheet('控制映射')
    ws.merge_cells('A1:I1')
    ws['A1'].value = 'MQTT 控制命令映射 (cloud/action/{SN}/control)'
    ws['A1'].font = Font(size=14, bold=True, color='1F3864')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:I2')
    ws['A2'].value = '消息格式: {"cmd_id":"<id>","value":<val>,"sn":"xxx","code":"xxx"}  |  ACK: cloud/action/{SN}/control/ack'
    ws['A2'].font = Font(size=9, color='666666')
    ws.row_dimensions[1].height = 28

    hdrs = ['cmd_id', '设备', '名称', 'data_dict键', '数据类型', '单位', '持久化', '处理器', '值示例']
    for ci, h in enumerate(hdrs, 1):
        ws.cell(row=4, column=ci, value=h).font = HF
        ws.cell(row=4, column=ci, value=h).fill = HFC; ws.cell(row=4, column=ci).alignment = HA; ws.cell(row=4, column=ci).border = BD
    ws.row_dimensions[4].height = 22

    sorted_ctrl = sorted(CTRL_MAP.items(), key=lambda x: int(x[0]))
    for ri, (cid, cfg) in enumerate(sorted_ctrl, start=1):
        row = ri + 4
        h = cfg.get('handler', '')
        dt = cfg.get('type', '-')
        device = cfg.get('device', 'ems')
        key = cfg.get('key', cfg.get('cmd', ''))
        display_name = key
        if cfg.get('cmd'):
            display_name = get_ctrl_zh(device, key)
        unit = EMS_DATA_DICT.get(key, {}).get('unit', '') if device == 'ems' and not h else ''
        example = {'bool': '0/1', 'int': '数值', 'float': '数值'}.get(dt, '')

        ws.cell(row, 1, int(cid)).alignment = CA; ws.cell(row, 2, device).alignment = CA
        ws.cell(row, 3, display_name).alignment = LA
        ws.cell(row, 4, key if not h else '-').alignment = LA
        ws.cell(row, 5, {'bool': 'BOOL', 'int': 'INT', 'float': 'FLOAT'}.get(dt, dt.upper())).alignment = CA
        ws.cell(row, 6, unit).alignment = CA
        ws.cell(row, 7, '是' if cfg.get('persist') else '否').alignment = CA
        if h == 'do':          ws.cell(row, 8, f'DO{cfg["num"]}开关').alignment = LA
        elif h == 'timing':    ws.cell(row, 8, '定时模式').alignment = LA
        elif h == 'demandResponse': ws.cell(row, 8, '需求响应').alignment = LA
        elif cfg.get('cmd'):   ws.cell(row, 8, '→ ct.control_dict').alignment = LA
        else:                  ws.cell(row, 8, '直接写入 data_dict').alignment = LA
        ws.cell(row, 9, example).alignment = CA
        for c in range(1, 10): ws.cell(row, c).border = BD
        if ri % 2 == 0:
            for c in range(1, 10): ws.cell(row, c).fill = AF

    for ci, w in enumerate([7, 6, 22, 22, 7, 6, 7, 18, 8], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'
    sw.cell(sr, 1, '控制映射').alignment = CA; sw.cell(sr, 2, '控制').alignment = CA
    sw.cell(sr, 3, '-').alignment = CA; sw.cell(sr, 4, '-').alignment = CA
    sw.cell(sr, 5, len(sorted_ctrl)).alignment = CA
    sw.cell(sr, 6, 'cmd_id 0-232, 订阅方通过 control 主题下发控制指令').alignment = LA
    for c in range(1, 7): sw.cell(sr, c).border = BD
    sr += 1

    # ===== 每设备发布数据表 =====
    for dev_name, zh_name, proto_path in PROTOCOL_DEVICES:
        offset = DEV_OFFSETS.get(dev_name, 0)
        is_ems = (dev_name == 'ems')
        registers = []

        if is_ems:
            # registers.append({'name': '在线状态', 'datatype': 'UINT16', 'unit': ''})
            for key, cfg in EMS_DATA_DICT.items():
                registers.append({
                    'name': key,
                    'datatype': cfg.get('datatype', 'UINT16'),
                    'unit': cfg.get('unit', ''),
                })
            registers.append({'name': '定时模式配置 (id=200)', 'datatype': 'JSON', 'unit': '', '_special': True})
            registers.append({'name': '需求响应配置 (id=300)', 'datatype': 'JSON', 'unit': '', '_special': True})
        elif proto_path:
            xml_regs = parse_xml(proto_path)
            # registers.append({'name': '在线状态', 'datatype': 'UINT16', 'unit': ''})
            registers.extend(xml_regs)
        else:
            continue

        # 分配 MQTT ID
        mid = offset
        for r in registers:
            r['mqtt_id'] = mid
            mid += 1

        ws = wb.create_sheet(f'{dev_name}_发布')
        ws.merge_cells('A1:F1')
        ws['A1'].value = f'设备: {zh_name} ({dev_name}) | MQTT ID 起始偏移: {offset}'
        ws['A1'].font = Font(size=14, bold=True, color='1F3864')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')
        ws['A2'].value = f'发布: cloud/push/{SN}/real | 格式: {{"id":<MQTT_ID>,"value":<val>}} | 增量(2s)+全量(60s)'
        ws['A2'].font = Font(size=9, color='666666')
        ws.row_dimensions[1].height = 28

        for ci, h in enumerate(['MQTT ID', '名称', '数据类型', '单位', '说明'], 1):
            ws.cell(row=4, column=ci, value=h).font = HF
            ws.cell(row=4, column=ci, value=h).fill = HFP; ws.cell(row=4, column=ci).alignment = HA; ws.cell(row=4, column=ci).border = BD
        ws.row_dimensions[4].height = 22

        for ri, r in enumerate(registers, start=1):
            row = ri + 4
            ws.cell(row, 1, r['mqtt_id']).alignment = CA
            ws.cell(row, 2, r['name']).alignment = LA
            ws.cell(row, 3, r['datatype']).alignment = CA
            ws.cell(row, 4, r['unit']).alignment = CA
            ws.cell(row, 5, '').alignment = LA
            for c in range(1, 6): ws.cell(row, c).border = BD
            if ri % 2 == 0:
                for c in range(1, 6): ws.cell(row, c).fill = AF

        for ci, w in enumerate([8, 34, 9, 6, 26], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A5'

        sw.cell(sr, 1, f'{dev_name}_发布').alignment = CA; sw.cell(sr, 2, '发布').alignment = CA
        sw.cell(sr, 3, dev_name).alignment = CA; sw.cell(sr, 4, zh_name).alignment = CA
        sw.cell(sr, 5, len(registers)).alignment = CA
        sw.cell(sr, 6, f'MQTT ID: {offset}~{mid-1}').alignment = LA
        for c in range(1, 7): sw.cell(sr, c).border = BD
        sr += 1
        print(f"  [{dev_name}] 发布: {len(registers)} 数据点 (offset={offset})")

    wb.save(output_path)
    print(f"\n✓ MQTT 协议文档: {os.path.abspath(output_path)}")


if __name__ == '__main__':
    generate_xlsx(sys.argv[1] if len(sys.argv) > 1 else 'mqtt_protocol.xlsx')
